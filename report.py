r"""
일마감 공정진척 보고서 생성 — **순수 openpyxl**. Windows/WSL 어디서나 돈다.

시트 구성
  1) 일마감요약   작업반별 50/80/모수/진행률 + 합계. 이 화면 하나로 끝나야 한다.
  2) 상태별집계   작업반 × 오더상태 교차표(작업지시 기준). 30이 어디에 몰렸는지 보인다.
  3) 점검         빠진 작업반·상태혼재·집계 오류. **문제를 숨기지 않는다.**
  4) 원본통합     전 작업반 원본 행 + 정규화한 오더상태코드 열
"""
from __future__ import annotations

import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import aggregate

TITLE_FONT = Font(name="맑은 고딕", size=14, bold=True)
HEAD_FONT = Font(name="맑은 고딕", size=10, bold=True, color="FFFFFF")
BODY_FONT = Font(name="맑은 고딕", size=10)
TOTAL_FONT = Font(name="맑은 고딕", size=10, bold=True)
SUB_FONT = Font(name="맑은 고딕", size=9, color="666666")

HEAD_FILL = PatternFill("solid", fgColor="365E8C")
TOTAL_FILL = PatternFill("solid", fgColor="DCE6F1")
WARN_FILL = PatternFill("solid", fgColor="FFF2CC")
BAD_FILL = PatternFill("solid", fgColor="F8CBAD")
GOOD_FILL = PatternFill("solid", fgColor="E2EFDA")

THIN = Side(style="thin", color="BFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")

# 진행률이 이 아래면 노랑, 더 아래면 주황으로 칠한다(눈에 띄게만 — 판정은 사람이 한다)
RATE_WARN = 0.90
RATE_BAD = 0.70


def _style_header(ws, row: int, ncol: int):
    for c in range(1, ncol + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEAD_FONT
        cell.fill = HEAD_FILL
        cell.alignment = CENTER
        cell.border = BOX


def _widths(ws, widths: list[int]):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _sheet_summary(wb, agg: dict, target_date: str, meta: dict):
    ws = wb.create_sheet("일마감요약")
    ws.sheet_view.showGridLines = False

    ws["A1"] = "공정진척현황(작업보고) 일마감 보고서"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:I1")
    ws["A2"] = (f"대상일자 {target_date}   |   iERP PM60250Rv3   |   "
                f"생성 {datetime.datetime.now():%Y-%m-%d %H:%M}")
    ws["A2"].font = SUB_FONT
    ws.merge_cells("A2:I2")
    ws["A3"] = ("집계기준 — 오더상태 30작업지시 제외 / 모수 = 50작업진행 + 80작업완료 / "
                "진행률 = 80작업완료 ÷ 모수 / 건수는 작업지시번호 기준")
    ws["A3"].font = SUB_FONT
    ws.merge_cells("A3:I3")

    head = ["작업반", "작업반명", "50작업진행", "80작업완료", "모수",
            "일마감 진행률", "30작업지시(제외)", "기타상태", "작업지시 계"]
    r0 = 5
    ws.append([])
    for i, h in enumerate(head, 1):
        ws.cell(row=r0, column=i, value=h)
    _style_header(ws, r0, len(head))

    r = r0 + 1
    for t in agg["teams"]:
        others = sum(t["others"].values())
        ws.cell(row=r, column=1, value=t["team"]).alignment = CENTER
        ws.cell(row=r, column=2, value=t["team_name"]).alignment = LEFT
        ws.cell(row=r, column=3, value=t["n50"])
        ws.cell(row=r, column=4, value=t["n80"])
        ws.cell(row=r, column=5, value=t["denom"])
        c = ws.cell(row=r, column=6, value=t["rate"])
        c.number_format = "0.0%"
        if t["rate"] is None:
            c.value = "—"
            c.alignment = CENTER
        elif t["rate"] < RATE_BAD:
            c.fill = BAD_FILL
        elif t["rate"] < RATE_WARN:
            c.fill = WARN_FILL
        else:
            c.fill = GOOD_FILL
        ws.cell(row=r, column=7, value=t["excluded_30"])
        ws.cell(row=r, column=8, value=others)
        ws.cell(row=r, column=9, value=t["order_total"])
        for cc in range(1, len(head) + 1):
            cell = ws.cell(row=r, column=cc)
            cell.border = BOX
            if cell.font is not BODY_FONT:
                cell.font = BODY_FONT
            if cc >= 3:
                cell.number_format = cell.number_format if cc == 6 else "#,##0"
                cell.alignment = CENTER
        r += 1

    tot = agg["total"]
    ws.cell(row=r, column=1, value="합계")
    ws.cell(row=r, column=3, value=tot["n50"])
    ws.cell(row=r, column=4, value=tot["n80"])
    ws.cell(row=r, column=5, value=tot["denom"])
    c = ws.cell(row=r, column=6, value=tot["rate"] if tot["rate"] is not None else "—")
    c.number_format = "0.0%"
    ws.cell(row=r, column=7, value=tot["excluded_30"])
    ws.cell(row=r, column=8, value=sum(sum(t["others"].values()) for t in agg["teams"]))
    ws.cell(row=r, column=9, value=tot["order_total"])
    for cc in range(1, len(head) + 1):
        cell = ws.cell(row=r, column=cc)
        cell.font = TOTAL_FONT
        cell.fill = TOTAL_FILL
        cell.border = BOX
        if cc >= 3:
            cell.alignment = CENTER
            if cc != 6:
                cell.number_format = "#,##0"
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)

    note = r + 2
    ws.cell(row=note, column=1,
            value="※ 진행률 색: 초록 ≥90% · 노랑 70~90% · 주황 <70% (표시일 뿐 판정 기준은 아님)"
            ).font = SUB_FONT
    if meta.get("failed"):
        ws.cell(row=note + 1, column=1,
                value=f"※ 수집 실패 작업반: {', '.join(meta['failed'])} — 모수에서 빠져 있습니다"
                ).font = Font(name="맑은 고딕", size=9, bold=True, color="C00000")

    _widths(ws, [10, 18, 12, 12, 10, 14, 16, 12, 12])
    ws.freeze_panes = ws.cell(row=r0 + 1, column=1)
    return ws


def _sheet_status(wb, agg: dict):
    ws = wb.create_sheet("상태별집계")
    ws.sheet_view.showGridLines = False
    codes = ["30", "50", "80", "55", "90", "20", "10", "00", "95"]
    present = [c for c in codes
               if any(t["order_counts"].get(aggregate.status_label(c)) for t in agg["teams"])]
    head = ["작업반", "작업반명"] + [aggregate.status_label(c) for c in present] + ["합계"]
    for i, h in enumerate(head, 1):
        ws.cell(row=1, column=i, value=h)
    _style_header(ws, 1, len(head))
    r = 2
    for t in agg["teams"]:
        ws.cell(row=r, column=1, value=t["team"]).alignment = CENTER
        ws.cell(row=r, column=2, value=t["team_name"])
        for j, c in enumerate(present, 3):
            ws.cell(row=r, column=j,
                    value=t["order_counts"].get(aggregate.status_label(c), 0))
        ws.cell(row=r, column=len(head), value=t["order_total"])
        for cc in range(1, len(head) + 1):
            cell = ws.cell(row=r, column=cc)
            cell.border = BOX
            cell.font = BODY_FONT
            if cc >= 3:
                cell.alignment = CENTER
                cell.number_format = "#,##0"
        r += 1
    ws.cell(row=r, column=1, value="합계").font = TOTAL_FONT
    for j, c in enumerate(present, 3):
        lbl = aggregate.status_label(c)
        ws.cell(row=r, column=j,
                value=sum(t["order_counts"].get(lbl, 0) for t in agg["teams"]))
    ws.cell(row=r, column=len(head), value=agg["total"]["order_total"])
    for cc in range(1, len(head) + 1):
        cell = ws.cell(row=r, column=cc)
        cell.font = TOTAL_FONT
        cell.fill = TOTAL_FILL
        cell.border = BOX
        if cc >= 3:
            cell.alignment = CENTER
            cell.number_format = "#,##0"
    _widths(ws, [10, 18] + [13] * len(present) + [10])
    ws.freeze_panes = "C2"
    return ws


def _sheet_check(wb, agg: dict, meta: dict):
    """★ 문제를 숨기지 않는 시트. 여기가 비어 있어야 보고서를 믿을 수 있다."""
    ws = wb.create_sheet("점검")
    ws.sheet_view.showGridLines = False
    ws["A1"] = "점검 — 아래가 비어 있어야 보고서를 그대로 믿을 수 있습니다"
    ws["A1"].font = TITLE_FONT
    r = 3
    for i, h in enumerate(["구분", "대상", "내용"], 1):
        ws.cell(row=r, column=i, value=h)
    _style_header(ws, r, 3)
    r += 1

    added = 0

    def add(kind, target, msg, warn=True):
        nonlocal r, added
        added += 1
        ws.cell(row=r, column=1, value=kind)
        ws.cell(row=r, column=2, value=target)
        ws.cell(row=r, column=3, value=msg)
        for cc in range(1, 4):
            cell = ws.cell(row=r, column=cc)
            cell.border = BOX
            cell.font = BODY_FONT
            cell.alignment = LEFT
            if warn:
                cell.fill = WARN_FILL
        r += 1

    for code in meta.get("failed", []):
        add("수집실패", code, "엑셀출력/저장에 실패해 이 작업반이 집계에서 빠졌습니다")
    for code in meta.get("missing", []):
        add("파일없음", code, "수집 폴더에 이 작업반 파일이 없습니다")
    for t in agg["teams"]:
        if t["order_total"] == 0:
            add("0건", t["team"], "조회 결과가 0건입니다 (그 날 실제로 없었는지 확인)")
        elif t["denom"] == 0:
            # ★ 진행률이 '—' 로 비는 이유를 반드시 남긴다. 조회는 됐는데 50/80 이 하나도
            #   없는 경우다(예: 전부 30작업지시). 숫자가 없는 것과 0% 인 것은 다르다.
            add("모수 0", t["team"],
                f"작업지시 {t['order_total']}건이 있으나 50작업진행·80작업완료가 0건이라 "
                f"진행률을 낼 수 없습니다 "
                f"(30작업지시 {t['excluded_30']}건"
                + (f", 기타 {sum(t['others'].values())}건" if t["others"] else "") + ")")
        if t["mixed"]:
            add("상태혼재", t["team"],
                f"한 작업지시에 오더상태가 2개 이상인 건 {t['mixed']}건 — "
                f"가장 진행된 상태로 셌습니다")
        if not t["has_order_col"]:
            add("열없음", t["team"],
                "'작업지시번호' 열을 못 찾아 행 단위로 셌습니다 — 건수가 부풀 수 있습니다")
    for e in agg["errors"]:
        add("집계오류", "", e)
    if added == 0:
        ws.cell(row=r, column=1, value="이상 없음").font = Font(
            name="맑은 고딕", size=10, bold=True, color="2E7D32")
        ws.cell(row=r, column=3, value="빠진 작업반·상태혼재·집계오류가 없습니다").font = SUB_FONT
    _widths(ws, [12, 14, 90])
    return ws


def _sheet_raw(wb, agg: dict):
    ws = wb.create_sheet("원본통합")
    header = None
    for t in agg["teams"]:
        if t["header"]:
            header = t["header"]
            break
    if not header:
        ws["A1"] = "원본 행 없음"
        return ws
    ws.append(["수집작업반", "오더상태코드"] + header)
    _style_header(ws, 1, len(header) + 2)
    for t in agg["teams"]:
        i_status = aggregate.find_col(t["header"], "status")
        for row in t["rows"]:
            code = aggregate.norm_status(row[i_status] if 0 <= i_status < len(row) else None)
            ws.append([t["team"], aggregate.status_label(code)] + list(row))
    _widths(ws, [12, 14] + [14] * len(header))
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = (f"A1:{get_column_letter(len(header) + 2)}{ws.max_row}")
    return ws


def build(agg: dict, target_date: str, out_path: Path, meta: dict | None = None) -> Path:
    """보고서 엑셀 1개를 만든다. 반환: 저장 경로."""
    meta = meta or {}
    wb = Workbook()
    wb.remove(wb.active)
    _sheet_summary(wb, agg, target_date, meta)
    _sheet_status(wb, agg)
    _sheet_check(wb, agg, meta)
    _sheet_raw(wb, agg)
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    try:
        wb.save(out_path)
    except PermissionError:
        # ★ 같은 이름의 보고서를 Excel 에서 열어 두면 잠긴다(재실행 때 흔하다).
        #   보고서를 통째로 잃는 대신 시각을 붙여 저장하고, 어디에 저장됐는지 알린다.
        import datetime as _dt
        alt = out_path.with_name(
            f"{out_path.stem}_{_dt.datetime.now():%H%M%S}{out_path.suffix}")
        wb.save(alt)
        print(f"※ {out_path.name} 이 열려 있어 저장하지 못했습니다 → {alt.name} 으로 저장했습니다")
        return alt
    return out_path


def build_from_folder(run_dir: Path, target_date: str, teams: list[dict] | None = None,
                      out_path: Path | None = None, meta: dict | None = None):
    agg = aggregate.aggregate_folder(run_dir, teams)
    out_path = out_path or Path(run_dir) / f"일마감_공정진척_{target_date.replace('-', '')}.xlsx"
    build(agg, target_date, out_path, meta)
    return out_path, agg
