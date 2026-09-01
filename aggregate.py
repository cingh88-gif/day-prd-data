r"""
공정진척현황(PM60250Rv3) 일마감 집계 — **순수 파싱**. Windows/WSL 어디서나 돈다.

★ 집계 규칙 (사용자 확정)
    · 오더상태 **30작업지시** 건은 **제외**한다.
    · 모수 = **50작업진행 건수 + 80작업완료 건수**
    · 일마감 진행률 = **80작업완료 / 모수**

★ 왜 화면에서 미리 30을 안 거르는가
  화면 조회조건으로 30을 빼 버리면 '30이 몇 건이었는지' 를 알 수 없다. 그러면 수집이
  덜 된 것인지, 원래 30이 많았던 것인지 구분이 안 된다. 그래서 **ALL 로 받아서 여기서**
  거르고, 제외한 건수까지 보고서에 남긴다.

★ '건' 의 단위
  공정진척현황은 한 작업지시가 공정별로 여러 행이 될 수 있다. 오더상태는 작업지시(오더)의
  속성이므로 같은 작업지시번호 행들은 상태가 같아야 한다.
  → **작업지시번호 기준 distinct** 를 정식 건수로 쓰고, 행 기준 건수도 같이 남긴다.
    한 작업지시에 상태가 섞여 있으면(=가정이 깨지면) 조용히 넘기지 않고 '혼재' 로 센다.
"""
from __future__ import annotations

import re
from collections import Counter, defaultdict
from pathlib import Path

# ── 오더상태 코드 (ierp-manhour 의 상태 콤보 실측값과 동일 체계) ──────────────
STATUS_NAMES = {
    "00": "오더아님", "10": "승인대기", "20": "오더확정", "30": "작업지시",
    "50": "작업진행", "55": "작업중단", "80": "작업완료", "90": "작업마감",
    "95": "작업취소",
}
NAME_TO_CODE = {v: k for k, v in STATUS_NAMES.items()}

EXCLUDE_CODES = {"30"}          # 모수에서 뺀다 (사용자 규칙)
DENOM_CODES = ["50", "80"]      # 모수 = 50 + 80
DONE_CODE = "80"                # 진행률 분자

# 한 작업지시에 상태가 섞였을 때 대표로 삼을 우선순위(진행이 더 된 쪽)
STATUS_PRIORITY = ["95", "00", "10", "20", "30", "50", "55", "80", "90"]

# iERP 가 끼워 넣는 합계 행 — 데이터가 아니다
SUBTOTAL_MARKERS = {"소계", "총계", "합계", "계"}

# ── 컬럼 별칭 — 실제 헤더가 무엇이든 여기에 걸리면 잡힌다 ────────────────────
#   ⚠️ 못 맞히면 **조용히 0을 채우지 않고** 실제 헤더를 보여주며 실패한다.
ALIASES = {
    "status": ["오더상태", "오더상태명", "작업상태", "진행상태", "상태", "지시상태"],
    "order_no": ["작업지시번호", "작업지시No", "작업지시NO", "지시번호", "오더번호",
                 "작업지시", "오더No", "제조오더번호"],
    # ⚠️ 작업반 ≠ 작업장. 섞으면 안 된다 (2026-09-01 실측: 이 화면의 엑셀에는 작업장 코드
    #   'C40001' 같은 값이 들어 있어, 작업장을 작업반으로 오인해 '작업반 불일치' 오탐이 났다).
    "team": ["작업반", "작업반코드", "작업반CD", "부서", "부서코드"],
    "team_name": ["작업반명", "부서명"],
    "workcenter": ["작업장", "작업장코드", "작업장명", "설비", "설비코드"],
    "item_cd": ["품목코드", "품번", "품목", "제품코드"],
    "item_nm": ["품목명", "품명", "제품명"],
    "qty_ord": ["지시수량", "계획수량", "오더수량", "작업지시수량"],
    "qty_done": ["실적수량", "생산수량", "완료수량", "보고수량", "양품수량"],
    "date": ["작업일자", "보고일자", "작업일", "지시일자", "일자", "완료일자"],
}


def norm_header(h) -> str:
    """헤더 비교용 정규화 — 공백/괄호/특수문자를 지운다."""
    return re.sub(r"[\s\(\)\[\]\/\.\-_·]", "", str(h or "")).strip()


def find_col(header: list[str], key: str) -> int:
    """별칭으로 열 위치를 찾는다. 없으면 -1."""
    normed = [norm_header(h) for h in header]
    for alias in ALIASES[key]:
        a = norm_header(alias)
        if a in normed:
            return normed.index(a)
    # 부분 일치 폴백 — '오더상태(코드)' 같은 변형을 잡는다
    for alias in ALIASES[key]:
        a = norm_header(alias)
        for i, h in enumerate(normed):
            if h and (h == a or h.startswith(a) or a.startswith(h) and len(h) >= 2):
                return i
    return -1


def norm_status(v) -> str | None:
    """'30작업지시' / '30' / '작업지시' → '30'. 못 알아보면 None."""
    s = str(v or "").strip()
    if not s:
        return None
    m = re.match(r"^\s*(\d{2})", s)
    if m and m.group(1) in STATUS_NAMES:
        return m.group(1)
    key = re.sub(r"\s", "", s)
    if key in NAME_TO_CODE:
        return NAME_TO_CODE[key]
    for name, code in NAME_TO_CODE.items():
        if name in key:
            return code
    return None


def status_label(code: str | None) -> str:
    if code is None:
        return "미상"
    return f"{code}{STATUS_NAMES.get(code, '')}"


def _to_num(v) -> float:
    try:
        return float(str(v).replace(",", "").strip())
    except (TypeError, ValueError):
        return 0.0


class HeaderError(RuntimeError):
    """필수 열을 못 찾았을 때 — 실제 헤더를 담아 던진다(조용히 0을 채우지 않는다)."""


def read_rows(xlsx_path: Path) -> tuple[list[str], list[list]]:
    """엑셀 1개 → (헤더, 데이터행들). 빈 행과 소계/총계 행은 걸러낸다."""
    from openpyxl import load_workbook
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    raw = list(ws.iter_rows(values_only=True))
    wb.close()
    if not raw:
        return [], []

    # iERP 가 제목 행을 먼저 깔기도 한다 → 별칭이 가장 많이 걸리는 행을 헤더로 본다.
    best_i, best_hits = 0, -1
    for i, r in enumerate(raw[:5]):
        hdr = [str(c).strip() if c is not None else "" for c in r]
        hits = sum(1 for k in ALIASES if find_col(hdr, k) >= 0)
        if hits > best_hits:
            best_i, best_hits = i, hits
    header = [str(c).strip() if c is not None else "" for c in raw[best_i]]

    rows = []
    for r in raw[best_i + 1:]:
        if all(v is None or str(v).strip() == "" for v in r):
            continue
        first_vals = [str(v).strip() for v in r[:3] if v is not None]
        if any(v in SUBTOTAL_MARKERS for v in first_vals):
            continue
        rows.append(list(r))
    return header, rows


def aggregate_team(xlsx_path: Path, team_code: str = "", team_name: str = "") -> dict:
    """엑셀 1개(작업반 1개분)를 집계한다."""
    header, rows = read_rows(xlsx_path)
    if not header:
        raise HeaderError(f"{xlsx_path.name}: 시트가 비어 있습니다")

    i_status = find_col(header, "status")
    if i_status < 0:
        raise HeaderError(
            f"{xlsx_path.name}: '오더상태' 열을 찾지 못했습니다.\n"
            f"  실제 헤더: {[h for h in header if h]}\n"
            f"  → aggregate.py 의 ALIASES['status'] 에 실제 열 이름을 추가하세요.")
    i_order = find_col(header, "order_no")
    i_team = find_col(header, "team")
    i_team_nm = find_col(header, "team_name")
    i_qty_ord = find_col(header, "qty_ord")
    i_qty_done = find_col(header, "qty_done")

    row_counts = Counter()                       # 상태별 행 수
    order_status = defaultdict(set)              # 작업지시번호 → 상태 집합
    qty_ord = defaultdict(float)
    qty_done = defaultdict(float)
    detected_name = team_name

    for n, r in enumerate(rows):
        def cell(i):
            return r[i] if 0 <= i < len(r) else None
        code = norm_status(cell(i_status))
        row_counts[code] += 1
        key = str(cell(i_order) or "").strip() if i_order >= 0 else f"__row{n}"
        if not key:
            key = f"__row{n}"
        order_status[key].add(code)
        if i_qty_ord >= 0:
            qty_ord[key] = max(qty_ord[key], _to_num(cell(i_qty_ord)))
        if i_qty_done >= 0:
            qty_done[key] += _to_num(cell(i_qty_done))
        if not detected_name and i_team_nm >= 0:
            detected_name = str(cell(i_team_nm) or "").strip()

    # 작업지시 단위로 대표 상태를 정한다
    order_counts = Counter()
    mixed = 0
    for key, codes in order_status.items():
        codes = {c for c in codes}
        if len(codes) > 1:
            mixed += 1
        rep = None
        for c in STATUS_PRIORITY:
            if c in codes:
                rep = c
        if rep is None:
            rep = next(iter(codes)) if codes else None
        order_counts[rep] += 1

    n50 = order_counts.get("50", 0)
    n80 = order_counts.get(DONE_CODE, 0)
    denom = n50 + n80
    excluded = sum(order_counts.get(c, 0) for c in EXCLUDE_CODES)
    others = {status_label(c): v for c, v in order_counts.items()
              if c not in EXCLUDE_CODES and c not in DENOM_CODES}

    return {
        "team": team_code or (str(rows[0][i_team]).strip()
                              if rows and i_team >= 0 and rows[0][i_team] else ""),
        "team_name": detected_name,
        "file": xlsx_path.name,
        "header": header,
        "rows": rows,
        "row_total": len(rows),
        "order_total": len(order_status),
        "n50": n50,
        "n80": n80,
        "denom": denom,
        "rate": (n80 / denom) if denom else None,
        "excluded_30": excluded,
        "others": others,
        "mixed": mixed,
        "row_counts": {status_label(c): v for c, v in row_counts.items()},
        "order_counts": {status_label(c): v for c, v in order_counts.items()},
        "qty_ord": sum(qty_ord.values()),
        "qty_done": sum(qty_done.values()),
        "has_order_col": i_order >= 0,
    }


def aggregate_folder(run_dir: Path, teams: list[dict] | None = None) -> dict:
    r"""수집 폴더(작업반별 xlsx 들) → 전체 집계 결과.

    파일명 규칙: progress_<작업반코드>_<YYYYMMDD>.xlsx
    """
    run_dir = Path(run_dir)
    name_map = {t["code"]: t.get("name", "") for t in (teams or [])}
    order = [t["code"] for t in (teams or [])]

    results, errors = [], []
    for f in sorted(run_dir.glob("progress_*.xlsx")):
        # progress_<작업반>_<기간라벨>.xlsx — 라벨은 20260831 / 202608 / 20260801-20260815
        m = re.match(r"progress_(.+?)_(\d{6,8}(?:-\d{8})?)\.xlsx$", f.name)
        code = m.group(1) if m else f.stem
        try:
            results.append(aggregate_team(f, code, name_map.get(code, "")))
        except HeaderError as e:
            errors.append(str(e))
        except Exception as e:
            errors.append(f"{f.name}: {e}")

    if order:
        rank = {c: i for i, c in enumerate(order)}
        results.sort(key=lambda r: rank.get(r["team"], 999))

    total = {
        "n50": sum(r["n50"] for r in results),
        "n80": sum(r["n80"] for r in results),
        "excluded_30": sum(r["excluded_30"] for r in results),
        "row_total": sum(r["row_total"] for r in results),
        "order_total": sum(r["order_total"] for r in results),
        "mixed": sum(r["mixed"] for r in results),
        "qty_ord": sum(r["qty_ord"] for r in results),
        "qty_done": sum(r["qty_done"] for r in results),
    }
    total["denom"] = total["n50"] + total["n80"]
    total["rate"] = (total["n80"] / total["denom"]) if total["denom"] else None

    return {"run_dir": run_dir, "teams": results, "total": total, "errors": errors}


def summary_text(agg: dict, target_date: str = "", kind: str = "일마감") -> str:
    """콘솔/GUI 용 한 눈 요약. kind = 일마감 / 월마감 / 기간."""
    t = agg["total"]
    lines = []
    if target_date:
        lines.append(f"■ {kind} · 개시예정일 {target_date}")
    lines.append(f"{'작업반':<14}{'50진행':>7}{'80완료':>7}{'모수':>7}{'진행률':>9}{'30제외':>8}")
    lines.append("-" * 54)
    for r in agg["teams"]:
        label = f"{r['team']} {r['team_name']}".strip()[:13]
        rate = f"{r['rate']:.1%}" if r["rate"] is not None else "—"
        lines.append(f"{label:<14}{r['n50']:>7,}{r['n80']:>7,}{r['denom']:>7,}"
                     f"{rate:>9}{r['excluded_30']:>8,}")
    lines.append("-" * 54)
    rate = f"{t['rate']:.1%}" if t["rate"] is not None else "—"
    lines.append(f"{'합계':<14}{t['n50']:>7,}{t['n80']:>7,}{t['denom']:>7,}"
                 f"{rate:>9}{t['excluded_30']:>8,}")
    if t["mixed"]:
        lines.append(f"※ 한 작업지시에 상태가 섞인 건 {t['mixed']}건 — 보고서 '점검' 시트 확인")
    for e in agg["errors"]:
        lines.append(f"※ {e}")
    return "\n".join(lines)
